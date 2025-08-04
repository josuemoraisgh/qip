import unicodedata
from sklearn.preprocessing import LabelEncoder
from openpyxl import load_workbook
from openpyxl.comments import Comment
import pandas as pd
from datetime import datetime

def diferenca_parcial(col1, formato_parcial_col1, col2, formato_col2):
    componentes = {'%Y': 'year', '%m': 'month', '%d': 'day','%H': 'hour', '%M': 'minute', '%S': 'second', '%f': 'microsecond'}
    def extrair_componentes(dt, formato, formato_ref):
        if isinstance(dt, pd.Timestamp):
            dt = dt.strftime(formato)
        dt_parse = datetime.strptime(dt, formato)
        comp_dt = {'year': 1900, 'month': 1, 'day': 1,'hour': 0, 'minute': 0, 'second': 0, 'microsecond': 0}
        for f, attr in componentes.items():
            if f in formato_ref:
                comp_dt[attr] = getattr(dt_parse, attr)
        return datetime(**comp_dt)

    diferencas = []
    for val1, val2 in zip(col1, col2):
        dt1 = extrair_componentes(val1, formato_parcial_col1, formato_parcial_col1)
        dt2 = extrair_componentes(val2, formato_col2, formato_parcial_col1)

        delta = (dt2 - dt1).total_seconds()
        diferencas.append(delta)

    return pd.Series(diferencas)

# Extrai a resposta esperada de uma coluna específica da aba "Pontuação"
def extrair_resposta_da_coluna(caminho_arquivo: str, nome_coluna: str, aba='Pontuação') -> str:
    wb = load_workbook(caminho_arquivo, data_only=True)
    if aba not in wb.sheetnames:
        return ""
    ws = wb[aba]
    col_index = None
    for col_num, cell in enumerate(ws[1], start=1):
        if cell.value and str(cell.value).strip() == nome_coluna.strip():
            col_index = col_num
            break
    if not col_index:
        return ""
    cell_val = ws.cell(row=2, column=col_index).value
    if not isinstance(cell_val, str):
        return ""
    if "R.:" in cell_val:
        return cell_val.split("R.:", 1)[1].strip()
    else:
        return ""

def substituir_resp(df: pd.DataFrame, cols_tela: list) -> pd.DataFrame:
    for column in cols_tela:
        cols = [col for col in df.columns if col.startswith(column) and col not in [f'{column}_part_1']]
        for col in cols:
            resp = extrair_resposta_da_coluna(file_path, col)
            # df[col] = df[col].map({resp:0}).fillna(1).astype('int')
            df[col] = (df[col] != resp).astype(int)
    return df

# Remove acentos e transforma o texto para minúsculo
def remover_acentos_e_transformar_minusculo(texto):
    texto = texto.lower()
    texto_sem_acentos = unicodedata.normalize('NFD', texto)
    texto_sem_acentos = texto_sem_acentos.encode('ascii', 'ignore').decode('utf-8')
    return texto_sem_acentos

# Marca 1 se o valor da coluna for string, 0 caso contrário (verifica presença de resposta)
def str_null(df: pd.DataFrame, cols_tela: list) -> pd.DataFrame:
    for column in cols_tela:
        cols = [col for col in df.columns if col.startswith(column) and col not in [f'{column}_part_1']]
        for column_name in cols:
            df[column_name] = df[column_name].apply(lambda x: 1 if isinstance(x, str) else 0)
    return df

# Converte respostas "Sim"/"Não" em valores binários (1/0)
def yes_no(df: pd.DataFrame, cols_tela: str) -> pd.DataFrame:
    for column in cols_tela:
        cols = [col for col in df.columns if col.startswith(column) and col not in [f'{column}_part_1']]
        for column_name in cols:
            df[column_name] = df[column_name].map({'Não': 0, 'Sim': 1, ' Não': 0, ' Sim': 1}).astype('int')
    return df
    
# Codifica colunas categóricas em números usando LabelEncoder
def label_encode_column(df: pd.DataFrame, cols_tela: list[str]) -> pd.DataFrame:
    for column in cols_tela:
        colunas = [col for col in df.columns if col.startswith(column) and col not in [f'{column}_part_1'] ]
        for coluna in colunas:
            if coluna in df.columns:
                le = LabelEncoder()
                df[coluna] = le.fit_transform(df[coluna].astype(str))
            else:
                print(f"Aviso: Coluna '{coluna}' não encontrada no DataFrame.")
    return df

# Converte string para datetime com formato esperado. Se falhar, retorna None
def extract_timestamp(date_str, col_name):
    try:
        return pd.to_datetime(date_str, format='%Y-%m-%d %H:%M:%S.%f')
    except Exception as e:
        print(f"Erro ao converter data na coluna '{col_name}': {date_str}")
        print(f"Motivo do erro: {e}")
        return None

# Insere comentários no cabeçalho de uma aba a partir de outra aba com as descrições
def inserir_comentarios_entre_abas(caminho_arquivo: str, sheet_destino: str, sheet_comentarios: str, linha_comentario: int = 2, pular_linhas_destino: int = 1):
    df_comentarios = pd.read_excel(caminho_arquivo, sheet_name=sheet_comentarios, skiprows=linha_comentario - 1, nrows=1, header=None)
    wb = load_workbook(caminho_arquivo)
    ws = wb[sheet_destino]
    for col_idx, comentario in enumerate(df_comentarios.iloc[0], start=1):
        if not pd.isna(comentario):
            ws.cell(row=1, column=col_idx).comment = Comment(str(comentario), "Resia Morais")
    wb.save(caminho_arquivo)

def classify_time(faixa: str, valor: int) -> int: # Criar a coluna de classificação
    if faixa == ' Mais que 1 hora' and valor > 70: return 0
    elif faixa == ' 5 minutos': 
        return 5 - valor
    elif faixa == ' 15 minutos': 
        return 15 - valor
    elif faixa == ' 30 minutos': 
        return 30 - valor
    elif faixa == ' 40 minutos': 
        return 40 - valor
    elif faixa == ' 60 minutos' or faixa == ' Mais que 1 hora': 
        return 60 - valor

if __name__ == '__main__':
    # Caminho para o arquivo xlsx local
    file_path = './python/banco_dados.xlsx'
    # Carregar o DataFrame
    df = pd.read_excel(file_path, sheet_name='XDados')        
    # Gerar a lista de colunas para aplicar a função
    columns_to_apply  = [f'Tela {i+1:02}_part_1' for i in range(77)]
    # Filtrar as colunas que realmente existem no DataFrame
    existing_columns = [col for col in columns_to_apply if col in df.columns]
    # Aplicar a função apenas nas colunas existentes
    #df[existing_columns] = df[existing_columns].map(extract_timestamp)
    for col in existing_columns:
        df[col] = df[col].apply(lambda x: extract_timestamp(x, col))        
    timeInicial = df['Tela 01_part_1']
    # Subtrair a próxima coluna pela coluna anterior
    for i in range(1, len(existing_columns)):
        df[existing_columns[i-1]] = pd.to_datetime(df[existing_columns[i]],format='%Y-%m-%d %H:%M:%S.%f',errors='coerce').sub(df[existing_columns[i-1]]).dt.total_seconds()
    df['Tela 77_part_1'] = pd.to_datetime(df['Tela 77_part_1'],format='%Y-%m-%d %H:%M:%S.%f',errors='coerce').sub(timeInicial).dt.total_seconds() / 60
    # Calcula o quanto o entrevistado estava errado em noção do dia em que ele esta
    df['Tela 02_part_2'] = (diferenca_parcial(df['Tela 02_part_2'],'%H:%M',timeInicial,'%Y-%m-%d %H:%M:%S.%f').abs() > 120).astype(int)  # se ele errou apenas em 3 minutos ou menos, considera-se que ele acertou o dia
    df['Tela 02_part_3'] = (diferenca_parcial(df['Tela 02_part_3'],'%d/%m/%Y',timeInicial,'%Y-%m-%d %H:%M:%S.%f').abs() > 86.400).astype(int)
    df = label_encode_column(df,[
        'Tela 02_part_5',  # Gênero
        'Tela 02_part_6',  # Sexo do nascimento 
        'Tela 02_part_7',  # Cor ou Raça
        'Tela 02_part_9',  # Estado civil
        'Tela 02_part_12', # Religião
        'Tela 02_part_13', # Escolaridade
        'Tela 02_part_14', # Renda familiar
        ])
    # Onde trem string coloca um e onde não tem string coloca zero
    cols_tela = ['Tela 07','Tela 10','Tela 43','Tela 53','Tela 57','Tela 62','Tela 63','Tela 64','Tela 65','Tela 66']
    df = str_null(df,cols_tela)
    # Onde trem string coloca um e onde não tem string coloca zero
    cols_tela = ['Tela 26','Tela 28','Tela 31','Tela 41','Tela 44','Tela 50','Tela 52','Tela 54','Tela 58','Tela 70','Tela 72']
    df = yes_no(df,cols_tela)
    # Extrai resposa do comentario e coloca zero na quela que for certa e um nas outras
    cols_tela = ['Tela 13','Tela 15','Tela 17','Tela 19','Tela 21','Tela 23','Tela 25', 'Tela 27','Tela 29','Tela 30','Tela 33']
    df = substituir_resp(df,cols_tela)
    # Tela 27: # Qual das imagens abaixo completa a sequência a seguir?
    cols_tela = ['Tela 27','Tela 30','Tela 33','Tela 47','Tela 48','Tela 49','Tela 59','Tela 60','Tela 69','Tela 71','Tela 73']        
    df = label_encode_column(df,cols_tela)                                                  
    # Tela 74: # 
    # FUTURO: Quando fazer aqui remove-lo da lista drop
    # Tela 75: # 
    # FUTURO: Quando fazer aqui remove-lo da lista drop
    # Tela 76:
    day_of_week = timeInicial.dt.day_name(locale='pt_BR').apply(remover_acentos_e_transformar_minusculo)
    df['Tela 76_part_2'] = df['Tela 76_part_2'].str.strip().apply(remover_acentos_e_transformar_minusculo)
    df['Tela 76_part_2'] = (df['Tela 76_part_2'] == day_of_week).astype(int)  
    # Tela 77: Calcula o quanto o entrevistado estava errado em noção do tempo gasto para resolver o questionário.
    df['Tela 77_part_2'] = df.apply(lambda row: classify_time(row['Tela 77_part_2'], row['Tela 77_part_1']), axis=1)    
    # Usando ExcelWriter para adicionar a nova aba ao arquivo Excel existente
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='TDados', index=False)
    inserir_comentarios_entre_abas(file_path,"TDados","Pontuação",2,0)