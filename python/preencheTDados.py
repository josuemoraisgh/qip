import unicodedata
from sklearn.preprocessing import LabelEncoder
from openpyxl import load_workbook
from openpyxl.comments import Comment
import pandas as pd
from datetime import datetime
import inspect

def normalizar_texto(texto: str) -> str:
    """Remove acentos e transforma em minúsculas."""
    texto = texto.lower()
    texto = unicodedata.normalize('NFD', texto)
    return texto.encode('ascii', 'ignore').decode('utf-8')

def expandir_opcoes_em_colunas(df: pd.DataFrame, coluna_base: str, opcoes: list) -> pd.DataFrame:
    """
    Substitui a coluna `{coluna_base}_part_2` por colunas binárias,
    com base na presença (via texto normalizado) das opções especificadas.
    Cada item de `opcoes` pode ser uma string ou uma lista de strings.
    """
    col_base = f"{coluna_base}_part_2"
    if col_base not in df.columns:
        raise ValueError(f"Coluna '{col_base}' não encontrada no DataFrame.")
    # Índice original da coluna a ser substituída
    col_base_idx = df.columns.get_loc(col_base)
    # Normaliza todos os valores da coluna base
    base_values = df[col_base].astype(str).apply(normalizar_texto)
    novas_colunas = {}
    for i, opcao in enumerate(opcoes):
        col_name = f"{coluna_base}_part_{i + 2}"
        if isinstance(opcao, str):
            termos = [normalizar_texto(opcao)]
        elif isinstance(opcao, list):
            termos = [normalizar_texto(term) for term in opcao if isinstance(term, str)]
        else:
            raise ValueError(f"Opção inválida no índice {i}: deve ser string ou lista de strings")
        # Verifica se algum dos termos normalizados está presente no valor da célula
        match = base_values.apply(lambda texto: any(termo in texto for termo in termos))
        novas_colunas[col_name] = match.astype(int)
    # Cria DataFrame com novas colunas
    df_novas = pd.DataFrame(novas_colunas, index=df.index)
    # Remove a coluna original e insere as novas no mesmo lugar
    df_sem_original = df.drop(columns=[col_base])
    df_esquerda = df_sem_original.iloc[:, :col_base_idx]
    df_direita = df_sem_original.iloc[:, col_base_idx:]
    df_final = pd.concat([df_esquerda, df_novas, df_direita], axis=1)
    return df_final

# Calcula a diferença entre dois conjuntos de datas parcialmente formatadas
# com base apenas nas partes especificadas no formato
def diferenca_parcial(col1, formato_parcial_col1, col2, formato_col2):
    componentes = {'%Y': 'year', '%m': 'month', '%d': 'day','%H': 'hour', '%M': 'minute', '%S': 'second', '%f': 'microsecond'}
    def extrair_componentes(dt, formato, formato_ref):
        if isinstance(dt, pd.Timestamp):
            dt = dt.strftime(formato)
        dt_parse = datetime.strptime(dt, formato)
        comp_dt = {'year': 1900, 'month': 1, 'day': 1,'hour': 0, 'minute': 0, 'second': 0, 'microsecond': 0}
        for f, attr in componentes.items():
            if f in formato_ref:
                comp_dt[attr] = getattr(dt_parse, attr)
        return datetime(**comp_dt)
    diferencas = []
    for val1, val2 in zip(col1, col2):
        dt1 = extrair_componentes(val1, formato_parcial_col1, formato_parcial_col1)
        dt2 = extrair_componentes(val2, formato_col2, formato_parcial_col1)
        delta = (dt2 - dt1).total_seconds()
        diferencas.append(delta)
    return pd.Series(diferencas)

# Extrai a resposta correta esperada da aba "Pontuação" de um arquivo Excel
def extrair_resposta_da_coluna(caminho_arquivo: str, nome_coluna: str, aba='Pontuação') -> str:
    try:
        df = pd.read_excel(caminho_arquivo, sheet_name=aba, nrows=2)
        if nome_coluna not in df.columns:
            return ""
        valor = df[nome_coluna].iloc[0]
        if not isinstance(valor, str):
            return ""
        return valor.split("R.:", 1)[1].strip() if "R.:" in valor else ""
    except Exception as e:
        print(f"[ERRO] Não foi possível extrair '{nome_coluna}' da aba '{aba}': {e}")
        return ""

# Aplica uma função personalizada para cada grupo de colunas por tela
# Agora a função trata dois casos:
# 1. Se a função recebida for um lambda simples (sem argumento), aplica diretamente
# 2. Se for um gerador de função (como gerar_transformador_resp), passa o nome da coluna como argumento
def aplicar_transformacao_personalizada(df: pd.DataFrame, colunas_por_tela: dict[str, callable]) -> pd.DataFrame:
    for tela, funcao in colunas_por_tela.items():
        colunas = [col for col in df.columns if col.startswith(tela) and col not in [f'{tela}_part_1']]
        for coluna in colunas:
            # Verifica se a função aceita apenas um argumento (a célula)
            if len(inspect.signature(funcao).parameters) == 1:
                df[coluna] = df[coluna].apply(funcao)
            else:
                # Assume-se que é uma função que precisa do nome da coluna
                df[coluna] = df[coluna].apply(lambda x: funcao(coluna,x))
    return df

# Codifica colunas categóricas em números
def label_encode_column(df: pd.DataFrame, cols_tela: list[str]) -> pd.DataFrame:
    for column in cols_tela:
        colunas = [col for col in df.columns if col.startswith(column) and col not in [f'{column}_part_1'] ]
        for coluna in colunas:
            if coluna in df.columns:
                le = LabelEncoder()
                df[coluna] = le.fit_transform(df[coluna].astype(str))
    return df

# Converte string para datetime, se falhar, retorna None
def extract_timestamp(date_str, col_name):
    try:
        return pd.to_datetime(date_str, format='%Y-%m-%d %H:%M:%S.%f')
    except Exception as e:
        print(f"Erro ao converter data na coluna '{col_name}': {date_str}")
        print(f"Motivo do erro: {e}")
        return None

# Insere comentários de uma aba de origem na primeira linha da aba de destino
def inserir_comentarios_entre_abas(caminho_arquivo: str, sheet_destino: str, sheet_comentarios: str, linha_comentario: int = 2, pular_linhas_destino: int = 1):
    df_comentarios = pd.read_excel(caminho_arquivo, sheet_name=sheet_comentarios, skiprows=linha_comentario - 1, nrows=1, header=None)
    wb = load_workbook(caminho_arquivo)
    ws = wb[sheet_destino]
    for col_idx, comentario in enumerate(df_comentarios.iloc[0], start=1):
        if not pd.isna(comentario):
            ws.cell(row=1, column=col_idx).comment = Comment(str(comentario), "Resia Morais")
    wb.save(caminho_arquivo)

# Classifica o erro de percepção de tempo com base na faixa estimada
def classify_time(faixa: str, valor: int) -> int:
    if faixa == ' Mais que 1 hora' and valor > 70: return 0
    if faixa.strip() in ['5 minutos', '15 minutos', '30 minutos', '40 minutos', '60 minutos']:
        return int(faixa.strip().split()[0]) - valor
    return 0

# Início do processamento principal
if __name__ == '__main__':
    file_path = './python/banco_dados.xlsx'
    df = pd.read_excel(file_path, sheet_name='XDados')        
    columns_to_apply  = [f'Tela {i+1:02}_part_1' for i in range(77)]
    existing_columns = [col for col in columns_to_apply if col in df.columns]

    # Converte strings para timestamps
    for col in existing_columns:
        df[col] = df[col].apply(lambda x: extract_timestamp(x, col))        
    timeInicial = df['Tela 01_part_1']

    # Calcula o tempo entre telas
    for i in range(1, len(existing_columns)):
        df[existing_columns[i-1]] = pd.to_datetime(df[existing_columns[i]],format='%Y-%m-%d %H:%M:%S.%f',errors='coerce').sub(df[existing_columns[i-1]]).dt.total_seconds()

    # Tempo total em minutos
    df['Tela 77_part_1'] = pd.to_datetime(df['Tela 77_part_1'],format='%Y-%m-%d %H:%M:%S.%f',errors='coerce').sub(timeInicial).dt.total_seconds() / 60

    # Verifica erro de noção de hora e data
    df['Tela 02_part_2'] = (diferenca_parcial(df['Tela 02_part_2'],'%H:%M',timeInicial,'%Y-%m-%d %H:%M:%S.%f').abs() > 120).astype(int)
    df['Tela 02_part_3'] = (diferenca_parcial(df['Tela 02_part_3'],'%d/%m/%Y',timeInicial,'%Y-%m-%d %H:%M:%S.%f').abs() > 86400).astype(int)

    # Label Encoding de dados demográficos
    df = label_encode_column(df,['Tela 02_part_5','Tela 02_part_6','Tela 02_part_7','Tela 02_part_9','Tela 02_part_12','Tela 02_part_13','Tela 02_part_14'])

    # Marca 1 se tiver texto nas colunas selecionadas
    df = aplicar_transformacao_personalizada(df, {
        tela : lambda x: int(isinstance(x, str))
        for tela in ['Tela 43','Tela 53','Tela 57','Tela 62','Tela 63','Tela 64','Tela 65','Tela 66']
    })

    # Converte respostas "Sim"/"Não" para 1/0
    df = aplicar_transformacao_personalizada(df, {
        tela: lambda x: {'Sim':1, 'Não':0, ' Sim':1, ' Não':0}.get(x, 0)
        for tela in ['Tela 26','Tela 28','Tela 31','Tela 41','Tela 44','Tela 50','Tela 52','Tela 54','Tela 58','Tela 70','Tela 72']
    })

    # Marca como incorreto (1) as respostas diferentes da correta que será zerada (0)
    df = aplicar_transformacao_personalizada(df, {
        tela: lambda nome_coluna, x: int(normalizar_texto(str(x)) != normalizar_texto(extrair_resposta_da_coluna(file_path, nome_coluna)))
        for tela in ['Tela 13','Tela 15','Tela 17','Tela 19','Tela 21','Tela 23','Tela 25','Tela 27','Tela 29','Tela 30','Tela 69']
    })

    # Codifica alternativas de múltipla escolha
    df = label_encode_column(df,['Tela 27','Tela 30','Tela 32','Tela 47','Tela 48','Tela 59','Tela 60','Tela 71','Tela 73'])

    df = expandir_opcoes_em_colunas(df,"Tela 33",['Manhã: 6:00 às 11:59 horas','Tarde: 12:00 às 17:59 horas','Noite: 18:00 às 23:59 horas','Madrugada: 00:00 às 05:59 horas'])
    df = expandir_opcoes_em_colunas(df,"Tela 42",['0',['1','2'],['3','4'],['5','6']])  
      
    # Verifica se o participante reconheceu corretamente o dia da semana
    day_of_week = timeInicial.dt.day_name(locale='pt_BR').apply(normalizar_texto)
    df['Tela 76_part_2'] = df['Tela 76_part_2'].str.strip().apply(normalizar_texto)
    df['Tela 76_part_2'] = (df['Tela 76_part_2'] == day_of_week).astype(int)

    # Avalia discrepância entre tempo declarado e tempo real
    df['Tela 77_part_2'] = df.apply(lambda row: classify_time(row['Tela 77_part_2'], row['Tela 77_part_1']), axis=1)

    # Salva o resultado na aba "TDados" e insere comentários de apoio
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='TDados', index=False)
    inserir_comentarios_entre_abas(file_path, "TDados", "Pontuação", 2, 0)
