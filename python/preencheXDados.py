import numpy as np
import pandas as pd
import unicodedata
from sklearn.preprocessing import OneHotEncoder
from openpyxl import load_workbook
from openpyxl.comments import Comment
        
# Funções de conversão
def remover_acentos_e_transformar_minusculo(texto):
    # Transforma o texto em minúsculas
    texto = texto.lower()
    # Normaliza a string para remover acentos
    texto_sem_acentos = unicodedata.normalize('NFD', texto)
    texto_sem_acentos = texto_sem_acentos.encode('ascii', 'ignore').decode('utf-8')
    return texto_sem_acentos

def substituir_string(texto, lista_substituicoes):  
    for padrao, novo in lista_substituicoes:
        if padrao.startswith('^'):
            if not padrao[1:] in texto: 
                texto = novo
        else:
            texto = texto.replace(padrao, novo)
    return texto
      
def str_null(df: pd.DataFrame, column: str) -> pd.DataFrame:
    cols_tela = [col for col in df.columns if col.startswith(column) and col not in [f'{column}_part_1'] ]    
    for column_name in cols_tela:
        df[column_name] = df[column_name].apply(lambda x: 1 if isinstance(x, str) else 0)        
    return df

def yes_no(df: pd.DataFrame, column: str) -> pd.DataFrame:
    cols_tela = [col for col in df.columns if col.startswith(column) and col not in [f'{column}_part_1'] ]    
    for column_name in cols_tela:
        df[column_name] = df[column_name].map({'Não': 0, 'Sim': 1, ' Não': 0, ' Sim': 1}).astype('int')
    return df

# Função para:
#   * remover espaços vazios antes e depois do ; e do |
#   * Dividir as colunas que têm valores separados por ; e por |
def split_columns(df):    
    columns_to_drop = []
    for col in df.columns:
        if df[col].dtype == 'object':
            # Remove espaços antes e depois de ; ou | usando regex
            df[col] = df[col].str.replace(r'\s*([;|])\s*', r'\1', regex=True)
            # Verifica se a coluna contém ; ou | e divide
            # Se a coluna contém ; ou |, divide em várias colunas
            if df[col].str.contains(r';|\|').any():
                split_df = df[col].str.split(r';|\|', expand=True)
                split_df.columns = [f'{col}_part_{i+1}' for i in range(split_df.shape[1])]
                df = pd.concat([df, split_df], axis=1)
                columns_to_drop.append(col)
    df = df.drop(columns=columns_to_drop)
    return df

# Função para transformar em vetor binário mantendo a data
def resposta_para_binario(valor):
    if not isinstance(valor, str) or ';' not in valor:
        return valor  # ignora casos inesperados

    partes = valor.split(';', 1)
    data = partes[0].strip()
    resposta = partes[1].strip()

    binarios = ['1' if resposta == opcao else '0' for opcao in opcoes]
    return f"{data};" + ';'.join(binarios) + ';'

# Insere comentários no cabeçalho da aba `sheet_destino` com base na linha `linha_comentario` da aba `sheet_comentarios`.
def inserir_comentarios_entre_abas(caminho_arquivo: str,sheet_destino: str,sheet_comentarios: str,linha_comentario: int = 2,pular_linhas_destino: int = 1):
    # Lê comentários da aba de origem
    df_comentarios = pd.read_excel(caminho_arquivo,sheet_name=sheet_comentarios,skiprows=linha_comentario - 1,nrows=1)
    # Lê os dados da aba de destino
    df_destino = pd.read_excel(caminho_arquivo,sheet_name=sheet_destino,skiprows=pular_linhas_destino)
    # Abre o arquivo Excel com openpyxl
    wb = load_workbook(caminho_arquivo)
    ws = wb[sheet_destino]
    # Reescreve os dados e insere comentários
    for col_idx, col_name in enumerate(df_destino.columns, start=1):
        # Escreve o cabeçalho original
        ws.cell(row=1, column=col_idx, value=col_name)
        # Insere comentário se existir na origem
        if col_name in df_comentarios.columns:
            comentario = str(df_comentarios[col_name].iloc[0])
            if pd.notna(comentario) and comentario.strip():
                ws.cell(row=1, column=col_idx).comment = Comment(comentario, "GPT")
        # Preenche os dados
        for row_idx, valor in enumerate(df_destino[col_name], start=2):
            ws.cell(row=row_idx, column=col_idx, value=valor)
    wb.save(caminho_arquivo)

if __name__ == '__main__':
    # Caminho para o arquivo xlsx local
    file_path = './python/banco_dados.xlsx'
    # Carregar o DataFrame
    df_inicial = pd.read_excel(file_path, sheet_name='BDados')
    # Realizando várias substituições em todas as colunas do DataFrame
    df_replace = df_inicial.replace({
        "other;": "",   # Remove "other;"
        "Sucess;": "",  # Remove "Sucess;"
        "Sucess": ""    # Remove "Sucess"
    }, regex=True)
    
    # Lista fixa de opções
    opcoes = ["Jesus Cristo","Coração","Dragão cuspindo fogo","Árvore","Não vi nada","Outra coisa"]
    # Aplicar às colunas desejadas
    for coluna in ['Tela 07', 'Tela 10']:
        if coluna in df_replace.columns:
            df_replace[coluna] = df_replace[coluna].apply(resposta_para_binario)

    # Tela 27:
    df_replace['Tela 27'] = (
        df_replace['Tela 27']
        .str.replace(r';{2,}', '', regex=True)  # Remove ";;"
        .str.replace(r'; ;', ';', regex=True)   # Remove "; ;" (espaço entre ;)
        .str.replace(r'; ', ';', regex=True)    # Remove espaço após ";"
    )
    # Tela 53:
    df_replace['Tela 53'] = (
        df_replace['Tela 53']
        .str.replace(r';{2,}', '', regex=True)  # Remove ";;"
        .str.replace(r'; ;', ';', regex=True)   # Remove "; ;" (espaço entre ;)
        .str.replace(r'; ', ';', regex=True)    # Remove espaço após ";"
    )
    # Aplicar a função para dividir as colunas
    df_split = split_columns(df_replace)
    # Remover as colunas vazias
    df_null = df_split.map(lambda x: np.nan if str(x).strip() == '' else x)
    df_dropna = df_null.dropna(axis=1, how='all')
    df = df_dropna.fillna(0)
    # Dentro de sua família, você é o(a) único(a) filho(a)?
    df['Tela 02_part_8'] = df['Tela 02_part_8'].replace("Sim", 0).astype('int')
    # Possui filhos(as)?
    df['Tela 02_part_10'] = df['Tela 02_part_10'].replace("Não", 0).astype('int')
    # Possui filhos(as) menores de 6 anos?
    df['Tela 02_part_11'] = df['Tela 02_part_11'].map({'Não': 0, 'Sim': 1}).astype('int')
    # Tela 42:
    def pontos_42(colData) -> int:
        index = 0.0
        for item in colData.split("|"):
            str = item.split(",")
            if 43.0 < float(str[0]) < 64.0 and 393.0 < float(str[1]) < 414.0: index += 1.0 #  53.799999999999955, 403.8 
            if 422.0 < float(str[0]) < 443.0 and 38.0 < float(str[1]) < 59.0: index += 1.0 # 432.79999999999995,   48.80000000000001;                
            if 28.0 < float(str[0]) < 49.0 and 201.0 < float(str[1]) < 222.0: index += 1.0 #  38.799999999999955, 211.8;                
            if 105.0 < float(str[0]) < 126.0 and 13.0 < float(str[1]) < 34.0: index += 1.0 # 115.79999999999995,   23.80000000000001;                \ \opkDF\
            if 274.0 < float(str[0]) < 295.0 and 369.0 < float(str[1]) < 390.0: index += 1.0 # 284.79999999999995,  379.8;                
            if 341.0 < float(str[0]) < 362.0 and 232.0 < float(str[1]) < 253.0: index += 1.0 # 351.79999999999995,  242.8                                                          
        return index
    df['Tela 42_part_2'] = df['Tela 42_part_2'].apply(pontos_42) 
    # Tela 51: # 
    def pontos_51(colData) -> int:
        index = 0.0
        for item in colData.split("|"):
            str = item.split(",")
            if 196.0 < float(str[0]) < 217.0 and 579.0 < float(str[1]) < 600.0: index += 1.0 # 206.79999999999995,589.8 
            if 105.0 < float(str[0]) < 126.0 and 585.0 < float(str[1]) < 606.0: index += 1.0 # 115.79999999999995,595.8   
            if 206.0 < float(str[0]) < 227.0 and 46.0 < float(str[1]) < 67.0: index += 1.0 # 216.79999999999995,56.80000000000001
            if 376.0 < float(str[0]) < 397.0 and 480.0 < float(str[1]) < 501: index += 1.0 # 386.79999999999995,490.8
            if 293.0 < float(str[0]) < 314.0 and 668.0 < float(str[1]) < 689.0: index += 1.0 # 303.79999999999995,678.3169811320754
        return index
    df['Tela 51_part_2'] = df['Tela 51_part_2'].apply(pontos_51)    
    df.insert(0,"Alvo", df['Tela 03_part_2'])
    df = df.drop(columns=['Tela 03_part_2', 'key','Mac Address_part_1','Mac Address_part_2','Mac Address_part_3'])
    # Usando ExcelWriter para adicionar a nova aba ao arquivo Excel existente
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='XDados', index=False)
    inserir_comentarios_entre_abas(file_path,"XDados","Pontuação",2,1)